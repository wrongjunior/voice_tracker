import openpyxl
from datetime import datetime
import os
import shutil


class SpreadsheetManager:
    def __init__(self, config: dict):
        self.file_path = config['excel_file_path']
        self.backup_folder = config.get('backup_folder', 'backups')
        self.max_backups = config.get('max_backups', 20)
        self.categories_map = config['categories']

        if not os.path.exists(self.file_path):
            raise FileNotFoundError(f"Файл Excel не найден по пути: {self.file_path}")

        os.makedirs(self.backup_folder, exist_ok=True)

    def _cleanup_old_backups(self):
        """
        Проверяет количество бэкапов и удаляет самые старые, если их больше лимита.
        """
        if self.max_backups <= 0:
            return

        try:
            all_files = os.listdir(self.backup_folder)
            backup_files = [f for f in all_files if f.startswith('backup_') and f.endswith('.xlsx')]

            if len(backup_files) <= self.max_backups:
                return

            backup_files.sort(key=lambda f: os.path.getctime(os.path.join(self.backup_folder, f)))
            files_to_delete = backup_files[:-self.max_backups]

            for filename in files_to_delete:
                file_path_to_delete = os.path.join(self.backup_folder, filename)
                os.remove(file_path_to_delete)
                print(f"Удален старый бэкап: {filename}")

        except OSError as e:
            print(f"Ошибка при очистке старых бэкапов: {e}")

    def _create_backup(self):
        """Создает резервную копию и запускает очистку старых."""
        timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        backup_file_path = os.path.join(self.backup_folder, f"backup_{timestamp}.xlsx")
        try:
            shutil.copy(self.file_path, backup_file_path)
            print(f"Резервная копия создана: {backup_file_path}")
            self._cleanup_old_backups()
        except Exception as e:
            print(f"Не удалось создать резервную копию: {e}")

    def _find_date_column(self, sheet):
        current_day = datetime.now().day
        for cell in sheet[1]:
            if cell.value == current_day:
                return cell.column
        return None

    def update_cell(self, category: str, value: int):
        self._create_backup()
        try:
            workbook = openpyxl.load_workbook(self.file_path)
            sheet = workbook.active
            row_num = self.categories_map.get(category)
            if not row_num:
                print(f"Ошибка: Категория '{category}' не найдена в конфигурации.")
                return False
            col_num = self._find_date_column(sheet)
            if not col_num:
                print(f"Ошибка: Столбец для дня '{datetime.now().day}' не найден.")
                return False
            sheet.cell(row=row_num, column=col_num, value=value)
            workbook.save(self.file_path)
            print(f"Успех! В категорию '{category}' добавлено значение '{value}'.")
            return True
        except Exception as e:
            print(f"Ошибка при работе с Excel: {e}")
            return False

    def get_stats_for_today(self):
        try:
            workbook = openpyxl.load_workbook(self.file_path)
            sheet = workbook.active
            col_num = self._find_date_column(sheet)
            if not col_num:
                return None, "Столбец для текущего дня не найден."
            stats = {}
            for category, row_num in self.categories_map.items():
                cell_value = sheet.cell(row=row_num, column=col_num).value
                stats[category] = cell_value if isinstance(cell_value, (int, float)) else 0
            return stats, None
        except Exception as e:
            return None, f"Ошибка при чтении статистики: {e}"