import openpyxl
from datetime import datetime
import os
import shutil


class SpreadsheetManager:
    def __init__(self, file_path, backup_folder, categories_map):
        self.file_path = file_path
        self.backup_folder = backup_folder
        self.categories_map = categories_map

        if not os.path.exists(self.file_path):
            raise FileNotFoundError(f"Файл Excel не найден по пути: {self.file_path}")

        # Создаем папку для бэкапов, если ее нет
        os.makedirs(self.backup_folder, exist_ok=True)

    def _create_backup(self):
        """Создает резервную копию файла с временной меткой."""
        timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        backup_file_path = os.path.join(self.backup_folder, f"backup_{timestamp}.xlsx")
        shutil.copy(self.file_path, backup_file_path)
        print(f"Резервная копия создана: {backup_file_path}")

    def _find_date_column(self, sheet):
        """Находит столбец, соответствующий текущему дню."""
        current_day = datetime.now().day
        header_row = sheet[1]  # Первая строка
        for cell in header_row:
            if cell.value == current_day:
                return cell.column
        return None

    def update_cell(self, category, value):
        """Обновляет ячейку для указанной категории и текущей даты."""
        self._create_backup()

        try:
            workbook = openpyxl.load_workbook(self.file_path)
            sheet = workbook.active

            # Находим строку по имени категории
            row_num = self.categories_map.get(category)
            if not row_num:
                print(f"Ошибка: Категория '{category}' не найдена в файле конфигурации.")
                return False

            # Находим столбец по текущей дате
            col_num = self._find_date_column(sheet)
            if not col_num:
                print(f"Ошибка: Столбец для сегодняшнего дня ({datetime.now().day}) не найден в файле.")
                return False

            # Записываем значение
            sheet.cell(row=row_num, column=col_num, value=value)

            workbook.save(self.file_path)
            print(f"Успех! В категорию '{category}' за {datetime.now().strftime('%d.%m.%Y')} добавлен балл.")
            return True

        except Exception as e:
            print(f"Произошла ошибка при работе с Excel: {e}")
            print("Восстановите файл из последней резервной копии в папке 'backups', если он повредился.")
            return False

    def get_stats_for_today(self):
        """Собирает статистику за текущий день и возвращает в виде словаря."""
        try:
            workbook = openpyxl.load_workbook(self.file_path)
            sheet = workbook.active

            col_num = self._find_date_column(sheet)
            if not col_num:
                print(f"Статистика: Столбец для сегодня не найден.")
                return None, "Столбец для текущего дня не найден в таблице."

            stats = {}
            # categories_map это {'обучение': 2, 'спорт': 6, ...}
            for category, row_num in self.categories_map.items():
                cell_value = sheet.cell(row=row_num, column=col_num).value
                # Если ячейка пустая или содержит не число, считаем за 0
                if isinstance(cell_value, (int, float)):
                    stats[category] = cell_value
                else:
                    stats[category] = 0

            return stats, None

        except Exception as e:
            error_message = f"Произошла ошибка при чтении статистики: {e}"
            print(error_message)
            return None, error_message